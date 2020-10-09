import openpyxl

from openpyxl import load_workbook

while True:
    print('Please enter the data you would like to add to your Excel document.')
    data = input()
    if (data == ''):
        print('Please enter some data.')
    else:
        while True:
            print('How is your data separated?')
            print('Please enter , (a comma) if your data is separated by commas. \n' +
            'Please enter ; (a semicolon) if your data is separated by semicolons. \n' +
            'Please enter  (a space) if your data is separated by spaces. \n')
            print('Note: Do not separate data with additional white space!')
            separationOptions = [',', ';', ' ']
            separationType = input()
            if separationType not in separationOptions:
                print('Please enter a valid separation method.')
            else:
                dataList = data.split(separationType)
                dataLength = len(dataList)
                while True:
                    print('Please enter the absolute path to an Excel document (e.g. Users/...)')
                    filePath = input()
                    if not (filePath.endswith('.xls') or filePath.endswith('.xlsx')):
                        print('Please input a valid Excel file.')
                    else:
                        try:
                            wb = load_workbook(filePath)
                            print('Please enter the name of the sheet where your data is located. \n' +
                            'Otherwise, the active sheet will be used by default.')
                            sheetName = input()
                            if (sheetName == ''):
                                ws = wb.active
                                print('You are using the active sheet.')
                            else:
                                try:
                                    ws = wb[sheetName]
                                    print(f'You are using the sheet: "{sheetName}"')
                                except KeyError:
                                    print('This sheet could not be found. The active sheet will be used instead.')
                                    ws = wb.active
                            while True:
                                print('Enter the letter of the column you would like to check for duplicates.')
                                columnLetter = input()
                                if (columnLetter == ''):
                                    print('Please enter a column letter.')
                                    continue
                                else:
                                    print('If you are counting these items in a separate column, ' +
                                    'please enter the corresponding letter for that column.\n' +
                                    'Otherwise, press ENTER to continue.')
                                    countColumn = input()
                                    sortedColumn = []
                                    rows = ws.max_row
                                    for i in range(rows):
                                        i += 1
                                        unsortedCell = ws[f'{columnLetter}{i}'].value
                                        sortedColumn.append(unsortedCell)
                                    currentRow = (int(rows) + 1)
                                    dataAdded = 0
                                    for i in range(dataLength):
                                        if dataList[i] in sortedColumn:
                                            print(f'{dataList[i]} already exists.')
                                            continue
                                        else:
                                            currentCell = (f'{columnLetter}{currentRow}')
                                            ws[currentCell] = dataList[i]
                                            if not ((countColumn != '') and (countColumn != columnLetter)):
                                                dataAdded += 1
                                                currentRow += 1
                                                continue
                                            else:
                                                previousRow = currentRow - 1
                                                previousCountCell = (f'{countColumn}{previousRow}')
                                                previousCount = ws[previousCountCell].value
                                                currentCountCell = (f'{countColumn}{currentRow}')
                                                ws[currentCountCell] = int(previousCount) + 1
                                                dataAdded += 1
                                                currentRow += 1
                                                continue
                                    print(f'Finished searching for duplicates. {dataAdded} value(s) added.')
                                    wb.save(filePath)
                                    print('Document saved!')
                                break
                            break
                        except IOError as e:
                            print(e.strerror)
                break
        break